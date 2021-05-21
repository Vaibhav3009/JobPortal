from django.shortcuts import render, redirect
from itertools import islice
import os
import openpyxl
from .models import JobPost,JobApply
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponseRedirect
import requests
from bs4 import BeautifulSoup
from jobs.jp import job_data
from django.views.decorators.csrf import csrf_exempt

# Create your views here.
# def home(request):
#     data = JobPost.objects.all()
#     # user_list = User.objects.all()
#     page = request.GET.get('page')

#     paginator = Paginator(data, 10)
#     try:
#         context = paginator.page(page)
#     except PageNotAnInteger:
#         context = paginator.page(1)
#     except EmptyPage:
#         context = paginator.page(paginator.num_pages)

#     # context={'data':data}
#     return render(request, 'jobs/index.html',{'context':context})

def home(request):
    return render(request, 'jobs/index.html')

def jobsresult(request):
 
 
    if request.method == 'POST':
        job = request.POST['jobp']
        naukri=job
        location = request.POST['jobl']
 
     
        job = job.strip()
        j = job.title().strip()
        items = j.split(' ')
        location = location.strip()
        
        job = job.replace(' ', '+')
 
        url ='https://www.indeed.co.in/jobs?q='+job+'&l='+location+'&start='
 
        id,x,y,z,a,b= job_data(url,items)
        
 
        data = list(zip(id,x,y,z,a,b))

        print(len(data))
        page = request.GET.get('page')

        paginator = Paginator(data, 10)
        try:
            dat = paginator.page(page)
        except PageNotAnInteger:
            dat = paginator.page(1)
        except EmptyPage:
            dat = paginator.page(paginator.num_pages)


        if len(x)>0:
            
            context ={
                'data':dat,
                'location':location,
                'job':naukri,
                 
            }
        
        else:
            context ={
                'message':'No Jobs Found..!',
                 
            }
 
 
        return render(request,'jobs/index.html',context)
     
    return render(request,'jobs/index.html')



def profile(request):
    user=request.user
    jobs=JobApply.objects.filter(user=user)
    # print(jobs)
    return render(request,'jobs/profile.html',{'jobs':jobs})
    # return render(request, 'jobs/about.html')

def about(request):
    return render(request, 'jobs/about.html')

def detail(request,pk):
    job = JobPost.objects.filter(id=pk).first()
    print(job)
    return render(request,'jobs/detail.html',{'job':job})

def apply(request,job_id):
    print(job_id)
    if request.method=="POST":
        user=request.user
        job=JobPost.objects.filter(id=job_id).first()
        JobApply(user=user,jobid=job).save()
        # jobs=JobApply.objects.filter(user=user)
        # print(jobs)
        # return render(request,'jobs/profile.html',{'jobs':jobs})
        return HttpResponseRedirect("/jobs/home/")


def profiledata(request):
    if request.method == 'POST' and request.is_ajax():
        user=request.user
        company = request.POST.get('company', False)
        location = request.POST.get('location', False)
        title = request.POST.get('title', False)
        link=request.POST.get('link', False)
        job=JobPost(company=company,joblocation_address=location,jobtitle=title,link=link)
        job.save()
        JobApply(user=user,jobid=job).save()

        return HttpResponse(status=201)
        


# def import_view(request):
#     cwd = os.getcwd()
#     file_to_read = cwd+'\\file\\sample.xlsx'
#     f = openpyxl.load_workbook(file_to_read)
#     sheet_obj = f.active
#     max_row = sheet_obj.max_row
#     max_col =sheet_obj.max_column
    
#     fList= []
#     for i in range(2, max_row):
#         templist = []
#         for j in range(1,max_col):
#             templist.append(sheet_obj.cell(row=i,column=j).value)
#         fList.append(templist)
   
#     batch_size = 1000
#     objs = (JobPost(company=data[0],education=data[1],experience=data[2],industry=data[3],jobdescription=data[4],jobid=data[5],joblocation_address=data[6],jobtitle=data[7],numberofpositions=data[8],payrate=data[9],postdate=data[10],site_name=data[11],skills=data[12],uniq_id=data[13]) for data in fList)
#     while True:
#         batch = list(islice(objs, batch_size))
#         if not batch:
#             break
#         JobPost.objects.bulk_create(batch, batch_size)
    
#     return HttpResponse('Success')
# import_view()
#         
